import openpyxl

inv_file = openpyxl.load_workbook("working with spread/inventory.xlsx")
product_list = inv_file["Sheet1"]


def get_each_company_with_product_count():
    product_per_supplier = {}
    for product_row in range(2, product_list.max_row + 1):
        supplier = product_list.cell(product_row, 4).value

        if supplier in product_per_supplier:
            product_per_supplier[supplier] = product_per_supplier.get(supplier) + 1
        else:
            product_per_supplier[supplier] = 1

    return product_per_supplier


def list_of_product_less_than_ten():
    products = []
    for product_row in range(2, product_list.max_row + 1):
        inventory_row = product_list.cell(product_row, 2).value

        if inventory_row < 10:
            products.append(product_list.cell(product_row, 1).value)

    return products


def total_value_of_inventory_in_the_product():
    product_per_supplier = {}
    for product_row in range(2, product_list.max_row + 1):
        supplier = product_list.cell(product_row, 4).value
        price = product_list.cell(product_row, 3).value
        inventory = product_list.cell(product_row, 2).value

        if supplier in product_per_supplier:
            product_per_supplier[supplier] = product_per_supplier[supplier] + price * inventory
        else:
            product_per_supplier[supplier] = inventory * price

    return product_per_supplier


# print(get_each_company_with_product_count())
print(list_of_product_less_than_ten())
# print(total_value_of_inventory_in_the_product())
